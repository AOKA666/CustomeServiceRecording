from flask import Flask, request, jsonify, Response
import requests
import json
import os
import time
import threading
from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
load_dotenv(os.path.join(BASE_DIR, '.env'))

app = Flask(__name__)
EXCEL_LOCK = threading.Lock()

CONFIG_FILE = os.path.join(os.path.dirname(__file__), 'config.json')
EXCEL_FILE = os.path.join(os.path.dirname(__file__), '..', 'info.xlsx')

EXCEL_HEADERS = ['来源', '项目', '电话', '微信', '省', '城市', '备注', '项目标签', '校区', '行为', 'row_id']


def check_auth():
    password = (os.environ.get("APP_PASSWORD") or "").strip()
    if not password:
        return True
    auth = request.authorization
    return bool(auth and auth.password == password)


@app.before_request
def require_auth():
    if check_auth():
        return None
    return Response(
        "需要访问密码",
        401,
        {"WWW-Authenticate": 'Basic realm="客服记录台"'}
    )

def reset_excel_file():
    wb = Workbook()
    ws = wb.active
    ws.append(EXCEL_HEADERS)
    wb.save(EXCEL_FILE)


def save_to_excel(row_data):
    with EXCEL_LOCK:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            first_row = [ws.cell(row=1, column=col).value for col in range(1, len(EXCEL_HEADERS) + 1)]
            if first_row != EXCEL_HEADERS:
                ws.insert_rows(1)
                for col, h in enumerate(EXCEL_HEADERS, 1):
                    ws.cell(row=1, column=col, value=h)
        else:
            reset_excel_file()
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

        row_values = [row_data.get(h, '') for h in EXCEL_HEADERS]
        ws.append(row_values)
        wb.save(EXCEL_FILE)
    return True

def load_config():
    defaults = {
        "yingdao_url": "http://127.0.0.1:9333/api/v1/robots/YOUR_ROBOT_ID/run",
        "cookie": ""
    }
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        for k, v in defaults.items():
            data.setdefault(k, v)
        return data
    return defaults

def load_project_options():
    path = os.path.join(BASE_DIR, '项目分类.txt')
    if not os.path.exists(path):
        return []
    with open(path, 'r', encoding='utf-8') as f:
        return [line.strip() for line in f if line.strip()]


def load_campus_data():
    path = os.path.join(BASE_DIR, '校区.txt')
    campus_options = []
    campus_map = {}
    if not os.path.exists(path):
        return campus_options, campus_map
    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split()
            if len(parts) >= 3:
                name = parts[0]
                province = parts[1]
                city = parts[2]
                campus_options.append(name)
                campus_map[name] = {"省": province, "城市": city}
    return campus_options, campus_map


def load_clue_data():
    path = os.path.join(BASE_DIR, '线索宝.txt')
    clue_options = []
    clue_map = {}
    if not os.path.exists(path):
        return clue_options, clue_map
    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split()
            if len(parts) >= 3:
                name = parts[0]
                province = parts[1]
                city = parts[2]
                clue_options.append(name)
                clue_map[name] = {"省": province, "城市": city}
    return clue_options, clue_map


def load_project_tags():
    path = os.path.join(BASE_DIR, '项目标签.txt')
    tag_map = {}
    if not os.path.exists(path):
        return tag_map
    current_project = None
    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            if line.startswith('- '):
                if current_project:
                    tag_map[current_project].append(line[2:].strip())
            else:
                current_project = line
                tag_map[current_project] = []
    return tag_map


def load_emergency_contacts():
    path = os.path.join(BASE_DIR, '紧急联系人.xlsx')
    if not os.path.exists(path):
        return {}
    try:
        wb = load_workbook(path)
        ws = wb.active
        contacts = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and row[1]:
                project = str(row[0]).strip()
                contact = str(row[1]).strip()
                contacts[project] = contact
        return contacts
    except Exception:
        return {}


PROVINCE_CITY_MAP = {
    "北京市": ["直辖区"],
    "天津市": ["直辖区"],
    "河北省": ["石家庄市", "唐山市", "秦皇岛市", "邯郸市", "邢台市", "保定市", "张家口市", "承德市", "沧州市", "廊坊市", "衡水市"],
    "山西省": ["太原市", "大同市", "阳泉市", "长治市", "晋城市", "朔州市", "晋中市", "运城市", "忻州市", "临汾市", "吕梁市"],
    "内蒙古自治区": ["呼和浩特市", "包头市", "乌海市", "赤峰市", "通辽市", "鄂尔多斯市", "呼伦贝尔市", "巴彦淖尔市", "乌兰察布市", "兴安盟", "锡林郭勒盟", "阿拉善盟"],
    "辽宁省": ["沈阳市", "大连市", "鞍山市", "抚顺市", "本溪市", "丹东市", "锦州市", "营口市", "阜新市", "辽阳市", "盘锦市", "铁岭市", "朝阳市", "葫芦岛市"],
    "吉林省": ["长春市", "吉林市", "四平市", "辽源市", "通化市", "白山市", "松原市", "白城市", "延边朝鲜族自治州"],
    "黑龙江省": ["哈尔滨市", "齐齐哈尔市", "鸡西市", "鹤岗市", "双鸭山市", "大庆市", "伊春市", "佳木斯市", "七台河市", "牡丹江市", "黑河市", "绥化市", "大兴安岭地区"],
    "上海市": ["直辖区"],
    "江苏省": ["南京市", "无锡市", "徐州市", "常州市", "苏州市", "南通市", "连云港市", "淮安市", "盐城市", "扬州市", "镇江市", "泰州市", "宿迁市"],
    "浙江省": ["杭州市", "宁波市", "温州市", "嘉兴市", "湖州市", "绍兴市", "金华市", "衢州市", "舟山市", "台州市", "丽水市"],
    "安徽省": ["合肥市", "芜湖市", "蚌埠市", "淮南市", "马鞍山市", "淮北市", "铜陵市", "安庆市", "黄山市", "滁州市", "阜阳市", "宿州市", "六安市", "亳州市", "池州市", "宣城市"],
    "福建省": ["福州市", "厦门市", "莆田市", "三明市", "泉州市", "漳州市", "南平市", "龙岩市", "宁德市"],
    "江西省": ["南昌市", "景德镇市", "萍乡市", "九江市", "新余市", "鹰潭市", "赣州市", "吉安市", "宜春市", "抚州市", "上饶市"],
    "山东省": ["济南市", "青岛市", "淄博市", "枣庄市", "东营市", "烟台市", "潍坊市", "济宁市", "泰安市", "威海市", "日照市", "临沂市", "德州市", "聊城市", "滨州市", "菏泽市"],
    "河南省": ["郑州市", "开封市", "洛阳市", "平顶山市", "安阳市", "鹤壁市", "新乡市", "焦作市", "濮阳市", "许昌市", "漯河市", "三门峡市", "南阳市", "商丘市", "信阳市", "周口市", "驻马店市"],
    "湖北省": ["武汉市", "黄石市", "十堰市", "宜昌市", "襄阳市", "鄂州市", "荆门市", "孝感市", "黄冈市", "咸宁市", "随州市", "恩施土家族苗族自治州"],
    "湖南省": ["长沙市", "株洲市", "湘潭市", "衡阳市", "邵阳市", "岳阳市", "常德市", "张家界市", "益阳市", "郴州市", "永州市", "怀化市", "娄底市", "湘西土家族苗族自治州"],
    "广东省": ["广州市", "韶关市", "深圳市", "珠海市", "汕头市", "佛山市", "江门市", "湛江市", "茂名市", "肇庆市", "惠州市", "梅州市", "汕尾市", "河源市", "阳江市", "清远市", "东莞市", "中山市", "潮州市", "揭阳市", "云浮市"],
    "广西壮族自治区": ["南宁市", "柳州市", "桂林市", "梧州市", "北海市", "防城港市", "钦州市", "贵港市", "玉林市", "百色市", "贺州市", "河池市", "来宾市", "崇左市"],
    "海南省": ["海口市", "三亚市", "三沙市", "儋州市", "五指山市", "琼海市", "文昌市", "万宁市", "东方市", "定安县", "屯昌县", "澄迈县", "临高县", "白沙黎族自治县", "昌江黎族自治县", "乐东黎族自治县", "陵水黎族自治县", "保亭黎族苗族自治县", "琼中黎族苗族自治县"],
    "重庆市": ["直辖区"],
    "四川省": ["成都市", "自贡市", "攀枝花市", "泸州市", "德阳市", "绵阳市", "广元市", "遂宁市", "内江市", "乐山市", "南充市", "眉山市", "宜宾市", "广安市", "达州市", "雅安市", "巴中市", "资阳市", "阿坝藏族羌族自治州", "甘孜藏族自治州", "凉山彝族自治州"],
    "贵州省": ["贵阳市", "六盘水市", "遵义市", "安顺市", "毕节市", "铜仁市", "黔西南布依族苗族自治州", "黔东南苗族侗族自治州", "黔南布依族苗族自治州"],
    "云南省": ["昆明市", "曲靖市", "玉溪市", "保山市", "昭通市", "丽江市", "普洱市", "临沧市", "楚雄彝族自治州", "红河哈尼族彝族自治州", "文山壮族苗族自治州", "西双版纳傣族自治州", "大理白族自治州", "德宏傣族景颇族自治州", "怒江傈僳族自治州", "迪庆藏族自治州"],
    "西藏自治区": ["拉萨市", "日喀则市", "昌都市", "林芝市", "山南市", "那曲市", "阿里地区"],
    "陕西省": ["西安市", "铜川市", "宝鸡市", "咸阳市", "渭南市", "延安市", "汉中市", "榆林市", "安康市", "商洛市"],
    "甘肃省": ["兰州市", "嘉峪关市", "金昌市", "白银市", "天水市", "武威市", "张掖市", "平凉市", "酒泉市", "庆阳市", "定西市", "陇南市", "临夏回族自治州", "甘南藏族自治州"],
    "青海省": ["西宁市", "海东市", "海北藏族自治州", "黄南藏族自治州", "海南藏族自治州", "果洛藏族自治州", "玉树藏族自治州", "海西蒙古族藏族自治州"],
    "宁夏回族自治区": ["银川市", "石嘴山市", "吴忠市", "固原市", "中卫市"],
    "新疆维吾尔自治区": ["乌鲁木齐市", "克拉玛依市", "吐鲁番市", "哈密市", "昌吉回族自治州", "博尔塔拉蒙古自治州", "巴音郭楞蒙古自治州", "阿克苏地区", "克孜勒苏柯尔克孜自治州", "喀什地区", "和田地区", "伊犁哈萨克自治州", "塔城地区", "阿勒泰地区", "石河子市", "阿拉尔市", "图木舒克市", "五家渠市", "北屯市", "铁门关市", "双河市", "可克达拉市", "昆玉市", "胡杨河市", "新星市"]
}


GAODUN_BASE_URL = "https://apigateway.gaodun.com"
GAODUN_SUBMIT_ENDPOINT = "/solon/api/v1/online-consultation/add-new"
GAODUN_REGIONS_ENDPOINT = "/prm/api/v1/regions/sub/get"
GAODUN_PROJECTS_ENDPOINT = "/solon/api/v1/authority/projects/without-conflict-projects"
GAODUN_CHANNELS_ENDPOINT = "/solon/api/v1/channel/online/get-channels"
GAODUN_SCOPE_ENDPOINT = "/solon/api/v1/online-consultation/get-online-user-scope"
GAODUN_QUICK_SEARCH_ENDPOINT = "/solon/api/v1/clues/global-quick-search"
DEFAULT_400_MARKET_CHANNEL_ID = 4583
DEFAULT_400_CAMPUS = "高顿网校 SEO"

PROJECT_TAG_RULES = {
    "公务员": {
        "kind": "scope_lookup",
        "result_index": 2,
        "match_by": "short_name",
        "fallback_name": "未知",
        "tag_template": {
            "tagCode": 12308048,
            "tagName": "省份",
            "checkStatus": True,
            "typeId": 12308,
            "singleChoice": 1,
        },
    },
    "银行招聘考试": {
        "kind": "scope_lookup",
        "result_index": 3,
        "match_by": "project_tag",
        "tag_template": {
            "tagCode": 1300001540,
            "tagName": "咨询意向-银行招聘考试",
            "checkStatus": True,
            "typeId": 12308,
            "singleChoice": 0,
        },
    },
    "学位教育": {
        "kind": "scope_lookup",
        "result_index": 0,
        "match_by": "project_tag",
        "tag_template": {
            "tagCode": 12308073,
            "tagName": "咨询意向-大职研",
            "checkStatus": True,
            "typeId": 12308,
            "singleChoice": 0,
        },
    },
    "CFA": {
        "kind": "scope_lookup",
        "result_index": 0,
        "match_by": "project_tag",
        "tag_template": {
            "tagCode": 12308034,
            "tagName": "咨询意向-金融",
            "checkStatus": True,
            "typeId": 12308,
            "singleChoice": 0,
        },
    },
    "证券从业": {
        "kind": "hardcoded",
        "tags_infos": [
            {
                "tagCode": 12308081,
                "tagName": "领取资料12月",
                "checkStatus": True,
                "typeId": 12308,
                "singleChoice": 1,
                "obTagsValueCheckedList": [
                    {"id": 4592, "objectId": None, "tagsId": 12308081,
                     "tagsValueCode": 4592, "tagsValueName": "证券从业意向"}
                ],
            }
        ],
    },
    "中级职称": {
        "kind": "hardcoded",
        "tags_infos": [
            {
                "tagCode": 12308153,
                "tagName": "中级通用-数据质量",
                "checkStatus": True,
                "typeId": 12308,
                "singleChoice": 1,
                "obTagsValueCheckedList": [
                    {"id": 4857, "objectId": None, "tagsId": 12308153,
                     "tagsValueCode": 4857, "tagsValueName": "正常数据-含未建联"}
                ],
            }
        ],
    },
    "CPA": {
        "kind": "hardcoded",
        "tags_infos": [
            {
                "tagCode": 12308154,
                "tagName": "CPA通用-数据质量",
                "checkStatus": True,
                "typeId": 12308,
                "singleChoice": 1,
                "obTagsValueCheckedList": [
                    {"id": 4867, "objectId": None, "tagsValueCode": 4867,
                     "tagsValueName": "正常数据-含未建联", "checkStatus": True}
                ],
            },
            {
                "tagCode": 1300037875,
                "tagName": "咨询意向-CPA",
                "checkStatus": True,
                "typeId": 12308,
                "singleChoice": 1,
                "obTagsValueCheckedList": [
                    {"id": 5461, "objectId": None, "tagsValueCode": 5461,
                     "tagsValueName": "高顿CPA意向", "checkStatus": True}
                ],
            },
        ],
    },
    "保研": {
        "kind": "scope_lookup",
        "result_index": 5,
        "match_by": "project_tag",
        "tag_template": {
            "tagCode": 1300077667,
            "tagName": "咨询意向-保研",
            "checkStatus": True,
            "typeId": 12308,
            "singleChoice": 1,
        },
    },
    "高级会计职称": {
        "kind": "hardcoded",
        "tags_infos": [
            {
                "tagCode": 1300076002,
                "tagName": "高会通用-数据质",
                "checkStatus": True,
                "typeId": 12308,
                "singleChoice": 1,
                "obTagsValueCheckedList": [
                    {"id": 5608, "objectId": None, "tagsId": 1300076002,
                     "tagsValueCode": 5608, "tagsValueName": "正常数据-含未建联"}
                ],
            }
        ],
    },
    "税务师": {
        "kind": "hardcoded",
        "tags_infos": [
            {
                "tagCode": 12308152,
                "tagName": "税务师通用-数据质量",
                "checkStatus": True,
                "typeId": 12308,
                "singleChoice": 1,
                "obTagsValueCheckedList": [
                    {"id": 4847, "objectId": None, "tagsId": 12308152,
                     "tagsValueCode": 4847, "tagsValueName": "正常数据-含未建联"}
                ],
            }
        ],
    },
    "中级经济师": {
        "kind": "hardcoded",
        "tags_infos": [
            {
                "tagCode": 1300000816,
                "tagName": "中级经济师通用-数据质",
                "checkStatus": True,
                "typeId": 12308,
                "singleChoice": 1,
                "obTagsValueCheckedList": [
                    {"id": 4948, "objectId": None, "tagsId": 1300000816,
                     "tagsValueCode": 4948, "tagsValueName": "正常数据-含未建联"}
                ],
            }
        ],
    },
}


def parse_cookie_string(cookie_str):
    cookies = {}
    if not cookie_str:
        return cookies
    for part in cookie_str.split(';'):
        part = part.strip()
        if not part or '=' not in part:
            continue
        k, _, v = part.partition('=')
        cookies[k.strip()] = v.strip()
    return cookies


def gaodun_headers(auth_token):
    return {
        "Accept": "application/json",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Authentication": auth_token,
        "Content-Type": "application/json; charset=UTF-8",
        "Origin": "https://ocrm.gaodun.com",
        "Referer": "https://ocrm.gaodun.com/",
        "Sec-Ch-Ua": '"Chromium";v="129", "Not=A?Brand";v="8"',
        "Sec-Ch-Ua-Mobile": "?0",
        "Sec-Ch-Ua-Platform": '"Windows"',
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/129.0.6668.101 Safari/537.36"
        ),
        "X-Requested-Extend": json.dumps({"systemName": "SYSTEM-OCRM"}),
    }


def _gaodun_get(path, params, auth_token, cookies):
    url = GAODUN_BASE_URL + path
    resp = requests.get(url, headers=gaodun_headers(auth_token),
                        params=params, cookies=cookies, timeout=15)
    resp.raise_for_status()
    return resp.json()


def _gaodun_post(path, payload, auth_token, cookies):
    url = GAODUN_BASE_URL + path
    resp = requests.post(url, headers=gaodun_headers(auth_token),
                         json=payload, cookies=cookies, timeout=15)
    resp.raise_for_status()
    return resp.json()


def lookup_teacher_name(mobile, project, config=None, attempts=4, delay_seconds=1.5):
    auth_token = (os.environ.get("GAODUN_AUTH_TOKEN") or "").strip()
    if not auth_token and config:
        auth_token = (config.get("auth_token") or "").strip()
    if not auth_token:
        return "查不到"

    payload = {
        "pageNum": 1,
        "pageSize": 10,
        "clueNo": "",
        "customerName": "",
        "customerNo": "",
        "mobile": mobile,
        "otherContact": "",
        "overseaMobile": "",
        "searchColumn": "mobile",
    }
    for attempt in range(1, attempts + 1):
        try:
            data = _gaodun_post(GAODUN_QUICK_SEARCH_ENDPOINT, payload, auth_token, {})
        except Exception as e:
            print(f"查询老师失败: {e}", flush=True)
            data = {}

        result_list = ((data.get("result") or {}).get("list") or [])
        matched = None
        if len(result_list) > 1:
            for item in result_list:
                if (item.get("intentProjectName") or "").strip() == project:
                    matched = item
                    break
        elif result_list:
            matched = result_list[0]

        if matched is not None:
            owner_name = matched.get("ownerName") or ""
            real_name = owner_name.split("-")[-1].strip()
            intent_project = (matched.get("intentProjectName") or project).strip()

            if not real_name or real_name == "新海" or real_name == "管理员":
                owner_name = load_emergency_contacts().get(intent_project, "查不到")
            return owner_name or "查不到"

        if attempt < attempts:
            print(f"老师查询未命中，等待后重试 {attempt}/{attempts}", flush=True)
            time.sleep(delay_seconds)

    return "查不到"


def lookup_state(state_name, auth_token, cookies):
    data = _gaodun_get(GAODUN_REGIONS_ENDPOINT,
                       {"parentId": 18784, "level": 1}, auth_token, cookies)
    for item in data.get("result") or []:
        if (item.get("name") or "").strip() == state_name:
            return item.get("id"), (item.get("shortName") or "").strip()
    raise ValueError(f"未找到省份 '{state_name}' (state_id)")


def lookup_city(city_name, state_id, auth_token, cookies):
    data = _gaodun_get(GAODUN_REGIONS_ENDPOINT,
                       {"parentId": state_id, "level": 2}, auth_token, cookies)
    for item in data.get("result") or []:
        if (item.get("name") or "").strip() == city_name:
            return item.get("id")
    raise ValueError(f"未找到城市 '{city_name}' (city_id)")


def lookup_project_id(project_name, auth_token, cookies):
    data = _gaodun_get(GAODUN_PROJECTS_ENDPOINT, {}, auth_token, cookies)
    for item in data.get("result") or []:
        if (item.get("name") or "").strip() == project_name:
            return item.get("id")
    raise ValueError(f"未找到项目 '{project_name}' (intent_project_id)")


def lookup_channel_id(channel_name, auth_token, cookies):
    if not channel_name:
        return None
    data = _gaodun_get(GAODUN_CHANNELS_ENDPOINT, {}, auth_token, cookies)
    items = data.get("result") or []
    candidates = [channel_name]
    if '-' in channel_name:
        candidates.append(channel_name.split('-', 1)[1])
    for cand in candidates:
        for item in items:
            if (item.get("name") or "").strip() == cand:
                return item.get("id")
    raise ValueError(f"未找到校区 '{channel_name}' (market_channel_id)")


def fetch_user_scope(intent_project_id, auth_token, cookies):
    payload = {"intentProjectIdList": [intent_project_id], "typeIdList": [12308]}
    return _gaodun_post(GAODUN_SCOPE_ENDPOINT, payload, auth_token, cookies)


def _find_scope_value(scope_data, result_index, target_name):
    results = scope_data.get("result") or []
    if result_index >= len(results):
        return None
    bucket = results[result_index] or {}
    for item in bucket.get("obTagsValueOutPutList") or []:
        if (item.get("tagsValueName") or "").strip() == target_name:
            return item
    return None


def build_tags_infos(project, project_tag, short_name, scope_data):
    rule = PROJECT_TAG_RULES.get(project)
    if rule is None:
        return None
    if rule.get("kind") == "hardcoded":
        return rule["tags_infos"]

    target_name = short_name if rule["match_by"] == "short_name" else project_tag
    found = _find_scope_value(scope_data, rule["result_index"], target_name)
    if found is None and rule.get("fallback_name"):
        found = _find_scope_value(scope_data, rule["result_index"], rule["fallback_name"])
    if found is None:
        raise ValueError(
            f"在 obTagsValueOutPutList 中未找到 tagsValueName='{target_name}'"
            f"{'（含 fallback 未知）' if rule.get('fallback_name') else ''}"
        )

    checked_item = {
        "id": found.get("id"),
        "objectId": found.get("objectId"),
        "tagsValueCode": found.get("tagsValueCode"),
        "tagsValueName": found.get("tagsValueName"),
        "checkStatus": True,
    }
    if found.get("tagsId") is not None:
        checked_item["tagsId"] = found.get("tagsId")

    tag = dict(rule["tag_template"])
    tag["obTagsValueCheckedList"] = [checked_item]
    return [tag]


def build_gaodun_payload(state, state_id, city, city_id, project, project_tag,
                         intent_project_id, market_channel_id, behavior_id,
                         mobile, follow_records, tags_infos):
    return {
        "beginTime": None,
        "behaviorId": behavior_id,
        "city": city,
        "cityId": city_id,
        "contactType": None,
        "countryId": 18784,
        "endTime": None,
        "followRecords": follow_records,
        "gender": 115003,
        "intentProjectId": intent_project_id,
        "lineTelPhone": "",
        "marketChannelId": market_channel_id,
        "mobile": mobile,
        "overSeaAreaCode": "",
        "professionId": 117004,
        "ruleType": "100521538",
        "specifiedDate": None,
        "state": state,
        "stateId": state_id,
        "tagsInfos": tags_infos or [],
        "trueName": "未留名",
    }


def normalize_row_data(row_data):
    normalized = dict(row_data or {})
    source = (normalized.get("来源") or "").strip()
    campus = (normalized.get("校区") or "").strip()
    if source == "400" and not campus:
        normalized["校区"] = DEFAULT_400_CAMPUS
    return normalized


def build_submit_display_text(row_data, follow_records, config=None):
    source = (row_data.get("来源") or "").strip()
    mobile = (row_data.get("电话") or "").strip()
    project = (row_data.get("项目") or "").strip()
    remark = (row_data.get("备注") or "").strip()
    if source != "400":
        return f"{mobile}, {follow_records}"

    print("表单提交成功，开始查询老师", flush=True)
    time.sleep(2)
    owner_name = lookup_teacher_name(mobile, project, config)
    real_name = owner_name.split("-")[-1].strip()
    user_name = ((config or {}).get("name") or "闫伟杰").strip()
    mobile2 = mobile[-4:] if len(mobile) >= 4 else mobile
    first_line = f"{owner_name}：老师这个学生咨询{project}，辛苦联系一下，电话：{mobile}，{remark}。"
    second_line = "\t".join([
        "接通",
        "呼入",
        mobile2,
        project,
        remark,
        "是",
        user_name,
        real_name,
    ])
    return first_line + "\n" + second_line


def submit_gaodun(row_data, config):
    auth_token = (os.environ.get("GAODUN_AUTH_TOKEN") or "").strip()
    if not auth_token:
        auth_token = (config.get("auth_token") or "").strip()
    cookie_text = (config.get("cookie") or "").strip()
    if not auth_token:
        raise ValueError("请先在 .env 中配置 GAODUN_AUTH_TOKEN")

    cookies = parse_cookie_string(cookie_text)
    source = (row_data.get("来源") or "").strip()
    project = (row_data.get("项目") or "").strip()
    project_tag = (row_data.get("项目标签") or "").strip()
    mobile = (row_data.get("电话") or "").strip()
    state = (row_data.get("省") or "").strip()
    city = (row_data.get("城市") or "").strip()
    campus = (row_data.get("校区") or "").strip()
    remark = (row_data.get("备注") or "").strip()

    state_id, short_name = lookup_state(state, auth_token, cookies)
    city_id = lookup_city(city, state_id, auth_token, cookies)
    intent_project_id = lookup_project_id(project, auth_token, cookies)
    if source == "400" and campus == DEFAULT_400_CAMPUS:
        market_channel_id = DEFAULT_400_MARKET_CHANNEL_ID
    else:
        market_channel_id = lookup_channel_id(campus, auth_token, cookies) if campus else None
    if market_channel_id is None and source == "400":
        campus = DEFAULT_400_CAMPUS
        market_channel_id = DEFAULT_400_MARKET_CHANNEL_ID
    behavior_id = 246 if source == "400" else 247
    follow_records = f"{project};{campus};{remark}"

    tags_infos = None
    if project in PROJECT_TAG_RULES:
        rule = PROJECT_TAG_RULES[project]
        scope_data = {}
        if rule.get("kind") == "scope_lookup":
            scope_data = fetch_user_scope(intent_project_id, auth_token, cookies)
        tags_infos = build_tags_infos(project, project_tag, short_name, scope_data)

    payload = build_gaodun_payload(
        state=state,
        state_id=state_id,
        city=city,
        city_id=city_id,
        project=project,
        project_tag=project_tag,
        intent_project_id=intent_project_id,
        market_channel_id=market_channel_id,
        behavior_id=behavior_id,
        mobile=mobile,
        follow_records=follow_records,
        tags_infos=tags_infos,
    )
    print("高顿提交 payload:", json.dumps(payload, ensure_ascii=False, indent=2), flush=True)
    response = _gaodun_post(GAODUN_SUBMIT_ENDPOINT, payload, auth_token, cookies)
    print("高顿提交 response:", json.dumps(response, ensure_ascii=False, indent=2), flush=True)
    status = response.get("status")
    code = response.get("code")
    if not (status == 0 or code == 0):
        raise ValueError(f"高顿接口返回失败：{json.dumps(response, ensure_ascii=False)}")
    return {
        "payload": payload,
        "response": response,
        "display_text": build_submit_display_text(row_data, follow_records, config),
    }


HTML = r"""<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>客服记录台</title>
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+SC:wght@300;400;500;700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

  :root {
    --bg: #f4f7fb;
    --surface: #ffffff;
    --surface2: #edf2fb;
    --border: #d8deed;
    --border-light: #c6cee3;
    --text: #1f2937;
    --text-muted: #5f6b85;
    --accent: #2f6feb;
    --accent-glow: rgba(47,111,235,0.12);
    --success: #34d399;
    --success-bg: rgba(52,211,153,0.1);
    --warning: #fbbf24;
    --danger: #f87171;
    --danger-bg: rgba(248,113,113,0.1);
    --radius: 8px;
  }

  html, body {
    height: 100%;
    background: var(--bg);
    color: var(--text);
    font-family: 'Noto Sans SC', sans-serif;
    font-size: 15px;
    line-height: 1.5;
  }

  body {
    display: flex;
    flex-direction: column;
    min-height: 100vh;
  }

  /* ── Header ── */
  header {
    background: var(--surface);
    border-bottom: 1px solid var(--border);
    padding: 0 16px;
    height: 48px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    position: sticky;
    top: 0;
    z-index: 100;
  }

  .logo {
    display: flex;
    align-items: center;
    gap: 10px;
    font-weight: 700;
    font-size: 17px;
    letter-spacing: 0.02em;
  }

  .logo-dot {
    width: 8px; height: 8px;
    border-radius: 50%;
    background: var(--accent);
    box-shadow: 0 0 10px var(--accent);
    animation: pulse 2s infinite;
  }

  @keyframes pulse {
    0%, 100% { opacity: 1; }
    50% { opacity: 0.4; }
  }

  .header-actions {
    display: flex;
    align-items: center;
    gap: 12px;
  }

  .counter {
    font-family: 'JetBrains Mono', monospace;
    font-size: 13px;
    color: var(--text-muted);
    background: var(--surface2);
    border: 1px solid var(--border);
    padding: 4px 10px;
    border-radius: 20px;
  }

  /* ── Main ── */
  main {
    flex: 1;
    padding: 16px 12px;
    overflow-x: auto;
  }

  /* ── Toolbar ── */
  .toolbar {
    display: flex;
    align-items: center;
    gap: 8px;
    margin-bottom: 12px;
  }

  .btn {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    padding: 8px 16px;
    border-radius: var(--radius);
    font-size: 14px;
    font-weight: 500;
    font-family: 'Noto Sans SC', sans-serif;
    cursor: pointer;
    border: none;
    transition: all 0.15s ease;
    white-space: nowrap;
  }

  .btn-primary {
    background: var(--accent);
    color: #fff;
  }
  .btn-primary:hover { background: #3a7de8; transform: translateY(-1px); }
  .btn-primary:active { transform: translateY(0); }

  .btn-ghost {
    background: transparent;
    color: var(--text-muted);
    border: 1px solid var(--border);
  }
  .btn-ghost:hover { background: var(--surface2); color: var(--text); border-color: var(--border-light); }

  .btn-danger-ghost {
    background: transparent;
    color: var(--danger);
    border: 1px solid transparent;
  }
  .btn-danger-ghost:hover { background: var(--danger-bg); border-color: var(--danger); }

  .btn-trigger {
    background: var(--surface2);
    color: var(--success);
    border: 1px solid rgba(52,211,153,0.25);
    padding: 4px 8px;
    font-size: 13px;
    border-radius: 5px;
  }
  .btn-trigger:hover {
    background: var(--success-bg);
    border-color: var(--success);
    transform: translateY(-1px);
  }
  .btn-trigger:active { transform: translateY(0); }
  .btn-trigger.loading {
    opacity: 0.6;
    cursor: not-allowed;
    pointer-events: none;
  }
  .btn-trigger.done {
    color: #64748b;
    border-color: transparent;
    background: transparent;
  }

  /* ── Table ── */
  .table-wrapper {
    border: 1px solid var(--border);
    border-radius: 10px;
    overflow: hidden;
    background: var(--surface);
  }

  table {
    width: 100%;
    border-collapse: collapse;
    min-width: 1100px;
  }

  thead tr {
    background: var(--surface2);
    border-bottom: 1px solid var(--border);
  }

  th {
    padding: 8px 6px;
    text-align: left;
    font-size: 13px;
    font-weight: 700;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    color: var(--text-muted);
    white-space: nowrap;
  }

  th.required::after {
    content: ' *';
    color: var(--danger);
  }

  tbody tr {
    border-bottom: 1px solid var(--border);
    transition: background 0.1s;
    animation: rowIn 0.2s ease;
  }

  @keyframes rowIn {
    from { opacity: 0; transform: translateY(-6px); }
    to   { opacity: 1; transform: translateY(0); }
  }

  tbody tr:last-child { border-bottom: none; }
  tbody tr:hover { background: var(--accent-glow); }
  tbody tr.completed-row { background: #dbeafe; }
  tbody tr.completed-row:hover { background: #c7dffb; }

  td {
    padding: 6px 5px;
    vertical-align: middle;
  }

  td[data-field] { position: relative; }

  .row-num {
    font-family: 'JetBrains Mono', monospace;
    font-size: 11px;
    color: var(--text-muted);
    text-align: center;
    width: 28px;
    padding: 0 2px;
  }

  /* ── Inputs ── */
  input[type="text"], textarea, select {
    width: 100%;
    background: #fff;
    border: 1px solid var(--border);
    border-radius: 5px;
    color: var(--text);
    font-family: 'Noto Sans SC', sans-serif;
    font-size: 13px;
    padding: 4px 6px;
    transition: border-color 0.15s, box-shadow 0.15s;
    outline: none;
    resize: none;
  }

  input[type="text"]:focus, textarea:focus, select:focus {
    border-color: var(--accent);
    box-shadow: 0 0 0 3px var(--accent-glow);
  }

  textarea { min-height: 48px; max-height: 80px; }

  td[data-field="备注"] textarea {
    padding-right: 18px;
  }

  .cell-toggle {
    position: absolute;
    top: 7px;
    right: 7px;
    width: 7px;
    height: 7px;
    padding: 0;
    border: 0;
    border-radius: 50%;
    background: currentColor;
    cursor: pointer;
    z-index: 2;
    opacity: 0.34;
    transition: opacity 0.15s, transform 0.15s, box-shadow 0.15s;
  }

  .cell-toggle:hover {
    opacity: 0.78;
    transform: scale(1.18);
  }

  .note-border-toggle {
    color: #ef4444;
  }

  .source-corner-toggle {
    color: #22c55e;
  }

  td.note-red-border textarea {
    border-color: #ef4444;
    box-shadow: 0 0 0 2px rgba(239,68,68,0.16);
  }

  td.note-red-border .note-border-toggle,
  td.source-corner .source-corner-toggle {
    opacity: 1;
    box-shadow: 0 0 0 2px rgba(255,255,255,0.95), 0 0 7px currentColor;
  }

  select option { background: var(--surface2); }

  /* ── Status badge ── */
  .status-badge {
    display: inline-flex;
    align-items: center;
    gap: 5px;
    font-size: 13px;
    padding: 3px 8px;
    border-radius: 20px;
    font-weight: 500;
  }
  .status-idle    { background: var(--surface2); color: var(--text-muted); }
  .status-ok      { background: var(--success-bg); color: var(--success); }
  .status-error   { background: var(--danger-bg); color: var(--danger); }
  .status-loading { background: rgba(79,142,247,0.1); color: var(--accent); }

  .status-dot {
    width: 5px; height: 5px;
    border-radius: 50%;
    background: currentColor;
  }

  /* ── Toast ── */
  #toast-container {
    position: fixed;
    bottom: 24px;
    right: 24px;
    display: flex;
    flex-direction: column;
    gap: 8px;
    z-index: 999;
  }

  .toast {
    display: flex;
    align-items: center;
    gap: 8px;
    background: var(--surface2);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 10px 16px;
    font-size: 14px;
    box-shadow: 0 8px 24px rgba(0,0,0,0.4);
    animation: toastIn 0.2s ease;
    min-width: 240px;
  }

  @keyframes toastIn {
    from { opacity: 0; transform: translateY(10px); }
    to   { opacity: 1; transform: translateY(0); }
  }

  .toast.success { border-color: rgba(52,211,153,0.3); }
  .toast.error   { border-color: rgba(248,113,113,0.3); }

  /* ── Empty state ── */
  .empty-state {
    text-align: center;
    padding: 60px 20px;
    color: var(--text-muted);
  }

  .empty-state .icon {
    font-size: 34px;
    margin-bottom: 12px;
    opacity: 0.4;
  }

  .empty-state p { font-size: 14px; }

  /* ── Column widths ── */
  .col-idx     { width: 30px; }
  .col-source  { width: 55px; }
  .col-proj    { width: 95px; }
  .col-phone   { width: 105px; }
  .col-wx      { width: 85px; }
  .col-province{ width: 55px; }
  .col-city    { width: 65px; }
  .col-note    { width: 150px; }
  .col-tag     { width: 95px; }
  .col-campus  { width: 125px; }
  .col-behavior{ width: 110px; }
  .col-status  { width: 65px; }
  .col-action  { width: 105px; }

  /* ── Result row ── */
  .result-row {
    background: var(--surface2);
    border-bottom: 1px solid var(--border);
  }
  .completed-row + .result-row {
    background: #d7e7fb;
  }
  .result-row td {
    padding: 6px 12px;
  }
  .result-row textarea {
    width: 100%;
    min-height: 56px;
    background: #fff;
    border: 1px solid var(--border);
    border-radius: 5px;
    font-family: 'JetBrains Mono', monospace;
    font-size: 13px;
    padding: 6px 8px;
    color: var(--text);
    resize: vertical;
  }
  .result-label {
    font-size: 11px;
    color: var(--text-muted);
    margin-bottom: 4px;
    letter-spacing: 0.05em;
  }

  /* ── Config modal ── */
  .modal-overlay {
    display: none;
    position: fixed; inset: 0;
    background: rgba(0,0,0,0.6);
    z-index: 500;
    align-items: center;
    justify-content: center;
  }
  .modal-overlay.show { display: flex; }

  .modal {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 28px;
    width: 480px;
    max-width: 95vw;
  }

  .modal h2 { font-size: 17px; margin-bottom: 6px; }
  .modal p  { color: var(--text-muted); font-size: 13px; margin-bottom: 20px; }

  .form-group { margin-bottom: 16px; }
  .form-group label { display: block; font-size: 13px; color: var(--text-muted); margin-bottom: 6px; }
  .form-group input, .form-group textarea { width: 100%; }
  .form-group textarea { min-height: 86px; max-height: 140px; resize: vertical; }

  .modal-footer { display: flex; gap: 8px; justify-content: flex-end; margin-top: 20px; }

  code {
    font-family: 'JetBrains Mono', monospace;
    font-size: 13px;
    background: var(--surface2);
    border: 1px solid var(--border);
    padding: 2px 6px;
    border-radius: 4px;
    color: var(--accent);
  }
</style>
</head>
<body>

<header>
  <div class="logo">
    <div class="logo-dot"></div>
    客服记录台
  </div>
  <div class="header-actions">
    <span class="counter" id="row-counter">0 条记录</span>
    <button class="btn btn-ghost" onclick="openConfig()">⚙ 配置</button>
  </div>
</header>

<main>
  <div class="toolbar">
    <button class="btn btn-primary" onclick="addRow()">
      <span>＋</span> 新增一行
    </button>
    <button class="btn btn-ghost" onclick="clearAll()">清空全部</button>
  </div>

  <div class="table-wrapper">
    <table id="main-table">
      <thead>
        <tr>
          <th class="col-idx">#</th>
          <th class="col-source required">来源</th>
          <th class="col-proj required">项目</th>
          <th class="col-phone required">电话</th>
          <th class="col-wx">微信</th>
          <th class="col-province required">省</th>
          <th class="col-city required">城市</th>
          <th class="col-note required">备注</th>
          <th class="col-tag">项目标签</th>
          <th class="col-campus">校区</th>
          <th class="col-behavior required">行为</th>
          <th class="col-status">状态</th>
          <th class="col-action">操作</th>
        </tr>
      </thead>
      <tbody id="table-body">
        <!-- rows injected here -->
      </tbody>
    </table>
    <div class="empty-state" id="empty-state">
      <div class="icon">📋</div>
      <p>点击「新增一行」开始记录通话信息</p>
    </div>
  </div>
</main>

<!-- Config Modal -->
<div class="modal-overlay" id="config-modal">
  <div class="modal">
    <h2>⚙ 提交配置</h2>
    <p>填写提交文案中使用的姓名。</p>
    <div class="form-group">
      <label>姓名</label>
      <input type="text" id="cfg-name" placeholder="闫伟杰">
    </div>
    <div class="modal-footer">
      <button class="btn btn-ghost" onclick="closeConfig()">取消</button>
      <button class="btn btn-primary" onclick="saveConfig()">保存</button>
    </div>
  </div>
</div>

<!-- Toast -->
<div id="toast-container"></div>

<script>
let PROJECT_OPTIONS = [];
let CAMPUS_OPTIONS = [];
let CAMPUS_MAP = {};
let CLUE_OPTIONS = [];
let CLUE_MAP = {};
let SOURCE_OPTIONS = [];
let PROJECT_TAG_MAP = {};
let PROVINCE_CITY_MAP = {};
let rowCount = 0;
let config = { name: '' };
const DEFAULT_400_CAMPUS = '高顿网校 SEO';

// ── Init ────────────────────────────────────────────────
function buildDatalists() {
  ['project-list', 'source-list', 'campus-list'].forEach(id => {
    const existing = document.getElementById(id);
    if (existing) existing.remove();
  });
  document.querySelectorAll('datalist[id^="tag-list-"]').forEach(el => el.remove());

  const dl = document.createElement('datalist');
  dl.id = 'project-list';
  PROJECT_OPTIONS.forEach(opt => {
    const option = document.createElement('option');
    option.value = opt;
    dl.appendChild(option);
  });
  document.body.appendChild(dl);

  const sourceDl = document.createElement('datalist');
  sourceDl.id = 'source-list';
  SOURCE_OPTIONS.forEach(opt => {
    const option = document.createElement('option');
    option.value = opt;
    sourceDl.appendChild(option);
  });
  document.body.appendChild(sourceDl);

  const campusDl = document.createElement('datalist');
  campusDl.id = 'campus-list';
  CAMPUS_OPTIONS.forEach(opt => {
    const option = document.createElement('option');
    option.value = opt;
    campusDl.appendChild(option);
  });
  CLUE_OPTIONS.forEach(opt => {
    const option = document.createElement('option');
    option.value = opt;
    campusDl.appendChild(option);
  });
  document.body.appendChild(campusDl);

  Object.entries(PROJECT_TAG_MAP).forEach(([project, tags]) => {
    const tagDl = document.createElement('datalist');
    tagDl.id = 'tag-list-' + project;
    tags.forEach(tag => {
      const option = document.createElement('option');
      option.value = tag;
      tagDl.appendChild(option);
    });
    document.body.appendChild(tagDl);
  });

  const provinceDl = document.createElement('datalist');
  provinceDl.id = 'province-list';
  Object.keys(PROVINCE_CITY_MAP).forEach(province => {
    const option = document.createElement('option');
    option.value = province;
    provinceDl.appendChild(option);
  });
  document.body.appendChild(provinceDl);
}

function buildCityDatalist(rowId, province) {
  const oldDl = document.getElementById('city-list-' + rowId);
  if (oldDl) oldDl.remove();
  const cities = province
    ? (PROVINCE_CITY_MAP[province] || [])
    : Array.from(new Set(Object.values(PROVINCE_CITY_MAP).flat()));
  if (cities.length === 0) return;
  const dl = document.createElement('datalist');
  dl.id = 'city-list-' + rowId;
  cities.forEach(city => {
    const option = document.createElement('option');
    option.value = city;
    dl.appendChild(option);
  });
  document.body.appendChild(dl);
}

function normalizeRegionName(value) {
  return (value || '').trim().replace(/\s+/g, '');
}

function stripCitySuffix(value) {
  return normalizeRegionName(value).replace(/(特别行政区|自治州|地区|盟|市|县|区)$/g, '');
}

function findProvinceByCity(cityName) {
  const normalizedCityName = normalizeRegionName(cityName);
  const shortCityName = stripCitySuffix(cityName);
  if (!normalizedCityName) return '';

  for (const [province, cities] of Object.entries(PROVINCE_CITY_MAP)) {
    for (const city of cities) {
      if (normalizeRegionName(city) === normalizedCityName || stripCitySuffix(city) === shortCityName) {
        return province;
      }
    }
  }
  return '';
}

function syncProvinceFromCity(cityInput) {
  const tr = cityInput.closest('tr');
  if (!tr) return false;

  const province = findProvinceByCity(cityInput.value);
  if (!province) return false;

  const provinceInput = tr.querySelector('td[data-field="省"] input[data-field="省"]');
  if (provinceInput && provinceInput.value !== province) {
    provinceInput.value = province;
    setFieldText(provinceInput, province);
  }
  buildCityDatalist(tr.dataset.id, province);
  return true;
}

window.onload = () => {
  loadConfig();
  fetch('/get_options')
    .then(r => r.json())
    .then(opts => {
      PROJECT_OPTIONS = opts.project_options || [];
      CAMPUS_OPTIONS = opts.campus_options || [];
      CAMPUS_MAP = opts.campus_map || {};
      CLUE_OPTIONS = opts.clue_options || [];
      CLUE_MAP = opts.clue_map || {};
      SOURCE_OPTIONS = opts.source_options || [];
      PROJECT_TAG_MAP = opts.project_tag_map || {};
      PROVINCE_CITY_MAP = opts.province_city_map || {};

      buildDatalists();

      document.getElementById('table-body').addEventListener('change', function(e) {
        const field = e.target.dataset.field;
        if (field === '项目') {
          updateTagField(e.target);
        } else if (field === '来源') {
          updateCampusFromSource(e.target);
        } else if (field === '校区') {
          onCampusChange(e.target);
        } else if (field === '省') {
          const tr = e.target.closest('tr');
          if (tr) buildCityDatalist(tr.dataset.id, e.target.value);
        } else if (field === '城市') {
          syncProvinceFromCity(e.target);
        }
        saveRowsToStorage();
      });

      document.getElementById('table-body').addEventListener('input', function(e) {
        if (e.target.dataset.field) {
          if (e.target.dataset.field === '城市') {
            syncProvinceFromCity(e.target);
          }
          saveRowsToStorage();
        }
      });

      document.getElementById('table-body').addEventListener('mousedown', function(e) {
        const input = e.target;
        if (input.tagName !== 'INPUT' || !input.hasAttribute('list')) return;
        if (input.value === '') return;
        const saved = input.value;
        input.value = '';
        input.dispatchEvent(new Event('input', { bubbles: true }));
        const restore = function() {
          if (input.value === '') {
            input.value = saved;
            input.dispatchEvent(new Event('input', { bubbles: true }));
          }
          input.removeEventListener('blur', restore);
        };
        input.addEventListener('blur', restore, { once: true });
      });

      restoreRows();
    })
    .catch(() => {
      toast('加载配置数据失败', 'error');
      restoreRows();
    });
};

// ── Config ──────────────────────────────────────────────
function loadConfig() {
  fetch('/get_config')
    .then(r => r.json())
    .then(d => {
      config = d;
      config.name = localStorage.getItem('crm_config_name') || d.name || '';
      document.getElementById('cfg-name').value = config.name || '';
    });
}

function openConfig() {
  document.getElementById('cfg-name').value = config.name || '';
  document.getElementById('config-modal').classList.add('show');
}

function closeConfig() {
  document.getElementById('config-modal').classList.remove('show');
}

function saveConfig() {
  const name = document.getElementById('cfg-name').value.trim();
  config.name = name;
  localStorage.setItem('crm_config_name', name);
  closeConfig();
  toast('配置已保存', 'success');
}

function createFieldElement(fieldName) {
  if (fieldName === '备注') {
    const textarea = document.createElement('textarea');
    textarea.placeholder = '备注内容...';
    textarea.dataset.field = fieldName;
    return textarea;
  }

  if (fieldName === '项目') {
    const input = document.createElement('input');
    input.type = 'text';
    input.setAttribute('list', 'project-list');
    input.placeholder = '输入或选择项目';
    input.dataset.field = fieldName;
    return input;
  }

  if (fieldName === '来源') {
    const input = document.createElement('input');
    input.type = 'text';
    input.setAttribute('list', 'source-list');
    input.placeholder = '输入或选择来源';
    input.dataset.field = fieldName;
    return input;
  }

  if (fieldName === '校区') {
    const input = document.createElement('input');
    input.type = 'text';
    input.setAttribute('list', 'campus-list');
    input.placeholder = '输入或选择校区';
    input.dataset.field = fieldName;
    return input;
  }

  if (fieldName === '行为') {
    const input = document.createElement('input');
    input.type = 'text';
    input.placeholder = '行为';
    input.dataset.field = fieldName;
    input.readOnly = true;
    input.value = '网络营销/来电';
    return input;
  }

  if (fieldName === '省') {
    const input = document.createElement('input');
    input.type = 'text';
    input.setAttribute('list', 'province-list');
    input.placeholder = '省';
    input.dataset.field = fieldName;
    return input;
  }

  const placeholders = {
    '电话': '手机号',
    '微信': '微信号',
    '省': '省',
    '城市': '城市',
    '项目标签': '标签',
    '校区': '校区'
  };

  const input = document.createElement('input');
  input.type = 'text';
  input.placeholder = placeholders[fieldName] || '';
  input.dataset.field = fieldName;
  return input;
}

function setFieldText(fieldEl, value) {
  const nextValue = value || '';
  fieldEl.value = nextValue;
  fieldEl.setAttribute('value', nextValue);
  fieldEl.title = nextValue;
  if (fieldEl.tagName === 'INPUT') {
    fieldEl.textContent = nextValue;
  }
}

function bindCellTextBridge(tr) {
  tr.querySelectorAll('td[data-field]').forEach(td => {
    const fieldName = td.dataset.field;
    let isSyncing = false;
    let boundFieldEl = null;
    let fieldAttrObserver = null;

    function onFieldInputOrChange() {
      if (isSyncing) return;
      isSyncing = true;
      const fieldEl = ensureElements();
      setFieldText(fieldEl, fieldEl.value || '');
      isSyncing = false;
    }

    function bindFieldEl(fieldEl) {
      if (boundFieldEl === fieldEl) return;

      if (boundFieldEl) {
        boundFieldEl.removeEventListener('input', onFieldInputOrChange);
        boundFieldEl.removeEventListener('change', onFieldInputOrChange);
      }
      if (fieldAttrObserver) {
        fieldAttrObserver.disconnect();
        fieldAttrObserver = null;
      }

      boundFieldEl = fieldEl;
      boundFieldEl.addEventListener('input', onFieldInputOrChange);
      boundFieldEl.addEventListener('change', onFieldInputOrChange);

      fieldAttrObserver = new MutationObserver((mutations) => {
        const touchedValue = mutations.some(m => m.type === 'attributes' && m.attributeName === 'value');
        if (!touchedValue) return;
        if (isSyncing) return;

        isSyncing = true;
        const attrValue = boundFieldEl.getAttribute('value') || '';
        if (boundFieldEl.value !== attrValue) {
          boundFieldEl.value = attrValue;
          boundFieldEl.title = attrValue;
          if (boundFieldEl.tagName === 'INPUT') {
            boundFieldEl.textContent = attrValue;
          }
          boundFieldEl.dispatchEvent(new Event('input', { bubbles: true }));
          boundFieldEl.dispatchEvent(new Event('change', { bubbles: true }));
        }
        isSyncing = false;
      });

      fieldAttrObserver.observe(boundFieldEl, {
        attributes: true,
        attributeFilter: ['value']
      });
    }

    function ensureElements() {
      let fieldEl = td.querySelector('[data-field]');
      const textInCell = td.textContent.trim();

      if (!fieldEl) {
        td.innerHTML = '';
        fieldEl = createFieldElement(fieldName);
        td.appendChild(fieldEl);
        setFieldText(fieldEl, textInCell);
      }

      bindFieldEl(fieldEl);
      return fieldEl;
    }

    function syncTdTextToField() {
      if (isSyncing) return;
      isSyncing = true;
      const fieldEl = ensureElements();
      const nextValue = td.textContent.trim();
      if (fieldEl.value !== nextValue) {
        setFieldText(fieldEl, nextValue);
        fieldEl.dispatchEvent(new Event('input', { bubbles: true }));
        fieldEl.dispatchEvent(new Event('change', { bubbles: true }));
      }
      isSyncing = false;
    }

    const initial = ensureElements();
    setFieldText(initial, initial.value || '');

    const observer = new MutationObserver(() => {
      const hasField = !!td.querySelector('[data-field]');
      if (!hasField) {
        ensureElements();
      }
      syncTdTextToField();
    });
    observer.observe(td, { childList: true, characterData: true, subtree: true });
  });
}

function setNoteBorder(id, enabled) {
  const tr = document.getElementById('row-' + id);
  const td = tr ? tr.querySelector('td[data-field="备注"]') : null;
  if (!td) return;
  td.classList.toggle('note-red-border', enabled);
  const btn = td.querySelector('.note-border-toggle');
  if (btn) btn.setAttribute('aria-pressed', enabled ? 'true' : 'false');
}

function toggleNoteBorder(id) {
  const tr = document.getElementById('row-' + id);
  const td = tr ? tr.querySelector('td[data-field="备注"]') : null;
  if (!td) return;
  setNoteBorder(id, !td.classList.contains('note-red-border'));
  saveRowsToStorage();
}

function setSourceCorner(id, enabled) {
  const tr = document.getElementById('row-' + id);
  const td = tr ? tr.querySelector('td[data-field="来源"]') : null;
  if (!td) return;
  td.classList.toggle('source-corner', enabled);
  const btn = td.querySelector('.source-corner-toggle');
  if (btn) btn.setAttribute('aria-pressed', enabled ? 'true' : 'false');
}

function toggleSourceCorner(id) {
  const tr = document.getElementById('row-' + id);
  const td = tr ? tr.querySelector('td[data-field="来源"]') : null;
  if (!td) return;
  setSourceCorner(id, !td.classList.contains('source-corner'));
  saveRowsToStorage();
}

function updateTagField(projectInput) {
  const tr = projectInput.closest('tr');
  if (!tr) return;
  const tagInput = tr.querySelector('td[data-field="项目标签"] input[data-field="项目标签"]');
  if (!tagInput) return;
  const project = projectInput.value.trim();
  const tags = PROJECT_TAG_MAP[project];
  if (tags && tags.length > 0) {
    tagInput.setAttribute('list', 'tag-list-' + project);
    tagInput.placeholder = '输入或选择标签';
  } else {
    tagInput.removeAttribute('list');
    tagInput.placeholder = '标签';
  }
}

function updateCampusFromSource(sourceInput) {
  const tr = sourceInput.closest('tr');
  if (!tr) return;
  const campusInput = tr.querySelector('td[data-field="校区"] input[data-field="校区"]');
  if (!campusInput) return;
  const source = sourceInput.value.trim();
  const raw = campusInput.dataset.rawCampus || '';
  if (raw) {
    if (CLUE_OPTIONS.includes(raw)) {
      campusInput.value = raw;
    } else {
      campusInput.value = source ? source + '-' + raw : raw;
    }
    setFieldText(campusInput, campusInput.value);
  }

  const behaviorInput = tr.querySelector('td[data-field="行为"] input[data-field="行为"]');
  if (behaviorInput) {
    const behavior = source === '400' ? '网络营销/400来电' : '网络营销/来电';
    behaviorInput.value = behavior;
    setFieldText(behaviorInput, behavior);
  }
}

function onCampusChange(campusInput) {
  const tr = campusInput.closest('tr');
  if (!tr) return;
  const sourceInput = tr.querySelector('td[data-field="来源"] input[data-field="来源"]');
  const source = sourceInput ? sourceInput.value.trim() : '';
  let raw = campusInput.value.trim();

  if (CLUE_OPTIONS.includes(raw)) {
    campusInput.dataset.rawCampus = raw;
    campusInput.value = raw;
    setFieldText(campusInput, raw);

    const clueMap = CLUE_MAP[raw];
    if (clueMap) {
      const provinceInput = tr.querySelector('td[data-field="省"] input[data-field="省"]');
      const cityInput = tr.querySelector('td[data-field="城市"] input[data-field="城市"]');
      if (provinceInput) {
        provinceInput.value = clueMap['省'] || '';
        setFieldText(provinceInput, provinceInput.value);
        buildCityDatalist(tr.dataset.id, provinceInput.value);
      }
      if (cityInput) {
        cityInput.value = clueMap['城市'] || '';
        setFieldText(cityInput, cityInput.value);
      }
    }
    return;
  }

  if (source && raw.startsWith(source + '-')) {
    raw = raw.slice(source.length + 1);
  }

  campusInput.dataset.rawCampus = raw;
  const finalValue = source ? source + '-' + raw : raw;
  campusInput.value = finalValue;
  setFieldText(campusInput, finalValue);

  const map = CAMPUS_MAP[raw];
  if (map) {
    const provinceInput = tr.querySelector('td[data-field="省"] input[data-field="省"]');
    const cityInput = tr.querySelector('td[data-field="城市"] input[data-field="城市"]');
    if (provinceInput) {
      provinceInput.value = map['省'] || '';
      setFieldText(provinceInput, provinceInput.value);
      buildCityDatalist(tr.dataset.id, provinceInput.value);
    }
    if (cityInput) {
      cityInput.value = map['城市'] || '';
      setFieldText(cityInput, cityInput.value);
    }
  }
}

// ── Table rows ──────────────────────────────────────────
function addRow() {
  rowCount++;
  const tbody = document.getElementById('table-body');
  const empty = document.getElementById('empty-state');
  empty.style.display = 'none';

  const id = Date.now();
  const tr = document.createElement('tr');
  tr.id = 'row-' + id;
  tr.dataset.id = id;

  tr.innerHTML = `
    <td class="row-num">${rowCount}</td>
    <td data-field="来源"><input type="text" list="source-list" placeholder="输入或选择来源" data-field="来源"><button type="button" class="cell-toggle source-corner-toggle" title="切换来源绿色圆点" aria-label="切换来源绿色圆点" aria-pressed="false" onclick="toggleSourceCorner(${id})"></button></td>
    <td data-field="项目"><input type="text" list="project-list" placeholder="输入或选择项目" data-field="项目"></td>
    <td data-field="电话"><input type="text" placeholder="手机号" data-field="电话"></td>
    <td data-field="微信"><input type="text" placeholder="微信号" data-field="微信"></td>
    <td data-field="省"><input type="text" list="province-list" placeholder="省" data-field="省"></td>
    <td data-field="城市"><input type="text" list="city-list-${id}" placeholder="城市" data-field="城市"></td>
    <td data-field="备注"><textarea placeholder="备注内容..." data-field="备注"></textarea><button type="button" class="cell-toggle note-border-toggle" title="切换备注红色边框" aria-label="切换备注红色边框" aria-pressed="false" onclick="toggleNoteBorder(${id})"></button></td>
    <td data-field="项目标签"><input type="text" placeholder="标签" data-field="项目标签"></td>
    <td data-field="校区"><input type="text" list="campus-list" placeholder="输入或选择校区" data-field="校区"></td>
    <td data-field="行为"><input type="text" placeholder="行为" data-field="行为" value="网络营销/来电" readonly></td>
    <td>
      <span class="status-badge status-idle" id="status-${id}">
        <span class="status-dot"></span>待触发
      </span>
    </td>
    <td style="padding:5px 4px;">
      <div style="display:flex; gap:4px; align-items:center; justify-content:center; height:100%;">
        <button class="btn btn-trigger" onclick="triggerRow(${id})" id="trigger-btn-${id}">
          ▶ 保存
        </button>
        <button class="btn btn-danger-ghost" onclick="deleteRow(${id})" style="padding:4px 6px; font-size:13px;">✕</button>
      </div>
    </td>
  `;

  tbody.appendChild(tr);

  const resultTr = document.createElement('tr');
  resultTr.id = 'result-row-' + id;
  resultTr.className = 'result-row';
  resultTr.dataset.resultFor = id;
  resultTr.innerHTML = `
    <td colspan="13">
      <div class="result-label">提交结果</div>
      <textarea id="result-text-${id}" readonly placeholder="提交完成后自动填写 mobile, follow_records"></textarea>
    </td>
  `;
  tbody.appendChild(resultTr);

  bindCellTextBridge(tr);
  buildCityDatalist(id, '');
  updateCounter();

  // Focus first input
  tr.querySelector('input').focus();
  saveRowsToStorage();
}

function deleteRow(id) {
  const tr = document.getElementById('row-' + id);
  const resultTr = document.getElementById('result-row-' + id);
  if (tr) {
    tr.style.animation = 'none';
    tr.style.opacity = '0';
    tr.style.transform = 'translateX(10px)';
    tr.style.transition = 'all 0.15s ease';
    if (resultTr) {
      resultTr.style.animation = 'none';
      resultTr.style.opacity = '0';
      resultTr.style.transition = 'all 0.15s ease';
    }
    setTimeout(() => {
      tr.remove();
      if (resultTr) resultTr.remove();
      updateCounter();
      if (document.getElementById('table-body').children.length === 0) {
        document.getElementById('empty-state').style.display = '';
      }
      renumberRows();
      saveRowsToStorage();
    }, 150);
  }
}

function renumberRows() {
  const rows = document.querySelectorAll('#table-body tr:not(.result-row)');
  rows.forEach((row, i) => {
    const numCell = row.querySelector('.row-num');
    if (numCell) numCell.textContent = i + 1;
  });
  rowCount = rows.length;
}

function clearAll() {
  if (!confirm('确认清空所有记录？')) return;
  document.getElementById('table-body').innerHTML = '';
  document.getElementById('empty-state').style.display = '';
  rowCount = 0;
  updateCounter();
  saveRowsToStorage();
}

// ── LocalStorage ────────────────────────────────────────
const STORAGE_KEY = 'crm_rows';
const STORAGE_TTL = 24 * 60 * 60 * 1000;

function saveRowsToStorage() {
  const rows = [];
  document.querySelectorAll('#table-body tr:not(.result-row)').forEach(tr => {
    const data = {};
    tr.querySelectorAll('input[data-field], textarea[data-field]').forEach(el => {
      data[el.dataset.field] = el.value;
    });
    const campusInput = tr.querySelector('[data-field="校区"]');
    if (campusInput) data._rawCampus = campusInput.dataset.rawCampus || '';
    const noteTd = tr.querySelector('td[data-field="备注"]');
    data._noteRedBorder = !!(noteTd && noteTd.classList.contains('note-red-border'));
    const sourceTd = tr.querySelector('td[data-field="来源"]');
    data._sourceCorner = !!(sourceTd && sourceTd.classList.contains('source-corner'));
    // persist completed state
    const btn = tr.querySelector('.btn-trigger');
    data._completed = btn && btn.classList.contains('done');
    const id = tr.dataset.id;
    const resultTa = document.getElementById('result-text-' + id);
    if (resultTa) data._result = resultTa.value;
    rows.push(data);
  });
  localStorage.setItem(STORAGE_KEY, JSON.stringify({ timestamp: Date.now(), rows }));
}

function restoreRows() {
  const raw = localStorage.getItem(STORAGE_KEY);
  if (!raw) return;
  try {
    const parsed = JSON.parse(raw);
    if (Date.now() - parsed.timestamp > STORAGE_TTL) {
      localStorage.removeItem(STORAGE_KEY);
      return;
    }
    parsed.rows.forEach(data => {
      if (Object.keys(data).length === 0) return;
      restoreRow(data);
    });
  } catch (e) {
    localStorage.removeItem(STORAGE_KEY);
  }
}

function restoreRow(data) {
  rowCount++;
  const tbody = document.getElementById('table-body');
  const empty = document.getElementById('empty-state');
  empty.style.display = 'none';

  const id = Date.now() + Math.floor(Math.random() * 1000);
  const tr = document.createElement('tr');
  tr.id = 'row-' + id;
  tr.dataset.id = id;

  tr.innerHTML = `
    <td class="row-num">${rowCount}</td>
    <td data-field="来源"><input type="text" list="source-list" placeholder="输入或选择来源" data-field="来源"><button type="button" class="cell-toggle source-corner-toggle" title="切换来源绿色圆点" aria-label="切换来源绿色圆点" aria-pressed="false" onclick="toggleSourceCorner(${id})"></button></td>
    <td data-field="项目"><input type="text" list="project-list" placeholder="输入或选择项目" data-field="项目"></td>
    <td data-field="电话"><input type="text" placeholder="手机号" data-field="电话"></td>
    <td data-field="微信"><input type="text" placeholder="微信号" data-field="微信"></td>
    <td data-field="省"><input type="text" list="province-list" placeholder="省" data-field="省"></td>
    <td data-field="城市"><input type="text" list="city-list-${id}" placeholder="城市" data-field="城市"></td>
    <td data-field="备注"><textarea placeholder="备注内容..." data-field="备注"></textarea><button type="button" class="cell-toggle note-border-toggle" title="切换备注红色边框" aria-label="切换备注红色边框" aria-pressed="false" onclick="toggleNoteBorder(${id})"></button></td>
    <td data-field="项目标签"><input type="text" placeholder="标签" data-field="项目标签"></td>
    <td data-field="校区"><input type="text" list="campus-list" placeholder="输入或选择校区" data-field="校区"></td>
    <td data-field="行为"><input type="text" placeholder="行为" data-field="行为" value="网络营销/来电" readonly></td>
    <td>
      <span class="status-badge status-idle" id="status-${id}">
        <span class="status-dot"></span>待触发
      </span>
    </td>
    <td style="padding:5px 4px;">
      <div style="display:flex; gap:4px; align-items:center; justify-content:center; height:100%;">
        <button class="btn btn-trigger" onclick="triggerRow(${id})" id="trigger-btn-${id}">
          ▶ 保存
        </button>
        <button class="btn btn-danger-ghost" onclick="deleteRow(${id})" style="padding:4px 6px; font-size:13px;">✕</button>
      </div>
    </td>
  `;

  tbody.appendChild(tr);

  const resultTr = document.createElement('tr');
  resultTr.id = 'result-row-' + id;
  resultTr.className = 'result-row';
  resultTr.dataset.resultFor = id;
  resultTr.innerHTML = `
    <td colspan="13">
      <div class="result-label">提交结果</div>
      <textarea id="result-text-${id}" readonly placeholder="提交完成后自动填写 mobile, follow_records"></textarea>
    </td>
  `;
  tbody.appendChild(resultTr);

  bindCellTextBridge(tr);
  updateCounter();

  const fields = ['来源', '项目', '电话', '微信', '省', '城市', '备注', '项目标签', '校区', '行为'];

  fields.forEach(field => {
    const td = tr.querySelector(`td[data-field="${field}"]`);
    const el = td ? td.querySelector('input[data-field], textarea[data-field]') : null;
    if (el && data[field] !== undefined) {
      el.value = data[field];
      setFieldText(el, data[field]);
    }
  });

  const campusTd = tr.querySelector('td[data-field="校区"]');
  const campusInput = campusTd ? campusTd.querySelector('input[data-field="校区"]') : null;
  if (campusInput && data._rawCampus) {
    campusInput.dataset.rawCampus = data._rawCampus;
  }

  const sourceTd = tr.querySelector('td[data-field="来源"]');
  const sourceInput = sourceTd ? sourceTd.querySelector('input[data-field="来源"]') : null;
  const projectTd = tr.querySelector('td[data-field="项目"]');
  const projectInput = projectTd ? projectTd.querySelector('input[data-field="项目"]') : null;
  if (sourceInput) updateCampusFromSource(sourceInput);
  if (projectInput) updateTagField(projectInput);
  setNoteBorder(id, !!data._noteRedBorder);
  setSourceCorner(id, !!data._sourceCorner);

  // restore city datalist for this row
  buildCityDatalist(id, data['省'] || '');

  // restore completed state
  if (data._completed) {
    tr.classList.add('completed-row');
    const btn = tr.querySelector('.btn-trigger');
    if (btn) {
      btn.classList.remove('loading');
      btn.classList.add('done');
      btn.textContent = '✓ 已完成';
    }
    setStatus(id, 'ok', '✓ 已保存');
  }

  // restore result text
  if (data._result !== undefined) {
    const resultTa = document.getElementById('result-text-' + id);
    if (resultTa) resultTa.value = data._result;
  }
}

function updateCounter() {
  const count = document.querySelectorAll('#table-body tr:not(.result-row)').length;
  document.getElementById('row-counter').textContent = count + ' 条记录';
}

// ── Trigger ─────────────────────────────────────────────
function getRowData(id) {
  const tr = document.getElementById('row-' + id);
  const data = {};
  tr.querySelectorAll('input[data-field], textarea[data-field], select[data-field]').forEach(el => {
    data[el.dataset.field] = el.value.trim();
  });
  return data;
}

function setStatus(id, type, text) {
  const el = document.getElementById('status-' + id);
  el.className = 'status-badge status-' + type;
  el.innerHTML = `<span class="status-dot"></span>${text}`;
}

function triggerRow(id) {
  const data = getRowData(id);

  // Validate required fields
  const required = ['来源', '项目', '电话', '省', '城市', '备注', '行为'];
  const missing = required.filter(f => !data[f] || !data[f].trim());
  if (missing.length > 0) {
    toast('请填写：' + missing.join('、'), 'error');
    return;
  }
  const tagRequiredProjects = Object.keys(PROJECT_TAG_MAP);
  if (tagRequiredProjects.includes(data['项目']) && (!data['项目标签'] || !data['项目标签'].trim())) {
    toast('请填写「项目标签」', 'error');
    return;
  }
  if (data['来源'] !== '400' && (!data['校区'] || !data['校区'].trim())) {
    toast('请填写「校区」（来源非400时必填）', 'error');
    return;
  }
  if (data['来源'] === '400' && (!data['校区'] || !data['校区'].trim())) {
    data['校区'] = DEFAULT_400_CAMPUS;
    const tr = document.getElementById('row-' + id);
    const campusInput = tr ? tr.querySelector('td[data-field="校区"] input[data-field="校区"]') : null;
    if (campusInput) {
      campusInput.value = DEFAULT_400_CAMPUS;
      setFieldText(campusInput, DEFAULT_400_CAMPUS);
    }
  }

  data['row_id'] = String(id);
  data['_config_name'] = config.name || '';

  const btn = document.getElementById('trigger-btn-' + id);
  const row = document.getElementById('row-' + id);
  if (row) row.classList.remove('completed-row');
  btn.classList.add('loading');
  btn.textContent = '⏳ 保存中...';
  setStatus(id, 'loading', '保存中');

  fetch('/save_excel', {
    method: 'POST',
    headers: { 'Content-Type': 'application/json' },
    body: JSON.stringify(data)
  })
  .then(r => r.json())
  .then(res => {
    if (res.success) {
      setStatus(id, 'ok', '✓ 已保存');
      btn.classList.remove('loading');
      btn.classList.add('done');
      btn.textContent = '✓ 已完成';
      if (row) row.classList.add('completed-row');
      toast(`第${getRowNum(id)}行 - 已保存到 Excel，并已提交表单`, 'success');
      const resultTa = document.getElementById('result-text-' + id);
      if (resultTa) resultTa.value = res.display_text || '';
      saveRowsToStorage();
    } else {
      throw new Error(res.error || '保存失败');
    }
  })
  .catch(err => {
    setStatus(id, 'error', '× 失败');
    btn.classList.remove('loading');
    btn.textContent = '↺ 重试';
    toast('保存失败: ' + err.message, 'error');
  });
}

function getRowNum(id) {
  const tr = document.getElementById('row-' + id);
  return tr ? tr.querySelector('.row-num').textContent : '?';
}

// ── Toast ────────────────────────────────────────────────
function toast(msg, type = 'info') {
  const c = document.getElementById('toast-container');
  const t = document.createElement('div');
  t.className = 'toast ' + type;
  t.innerHTML = (type === 'success' ? '✓ ' : type === 'error' ? '✕ ' : '') + msg;
  c.appendChild(t);
  setTimeout(() => {
    t.style.opacity = '0';
    t.style.transform = 'translateY(10px)';
    t.style.transition = 'all 0.2s';
    setTimeout(() => t.remove(), 200);
  }, 2800);
}

// Close modal on overlay click
document.getElementById('config-modal').addEventListener('click', function(e) {
  if (e.target === this) closeConfig();
});
</script>
</body>
</html>"""


@app.route('/')
def index():
    return HTML


@app.route('/get_config')
def get_config():
    return jsonify(load_config())


@app.route('/get_options')
def get_options():
    campus_options, campus_map = load_campus_data()
    clue_options, clue_map = load_clue_data()
    return jsonify({
        "project_options": load_project_options(),
        "campus_options": campus_options,
        "campus_map": campus_map,
        "clue_options": clue_options,
        "clue_map": clue_map,
        "project_tag_map": load_project_tags(),
        "emergency_contacts": load_emergency_contacts(),
        "source_options": ['400', '美团', '知了', '高德'],
        "province_city_map": PROVINCE_CITY_MAP
    })


@app.route('/save_config', methods=['POST'])
def save_config():
    data = load_config()
    data.update(request.json or {})
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return jsonify({"success": True})


@app.route('/trigger', methods=['POST'])
def trigger():
    row_data = request.json
    config = load_config()
    url = config.get('yingdao_url', '').strip()

    if not url or 'YOUR_ROBOT_ID' in url:
        return jsonify({
            "success": False,
            "error": "请先在右上角「⚙ 配置」中填写影刀触发URL"
        }), 400

    try:
        resp = requests.post(
            url,
            json={"params": row_data},
            timeout=15,
            headers={"Content-Type": "application/json"}
        )
        resp.raise_for_status()
        return jsonify({"success": True, "response": resp.text})
    except requests.exceptions.ConnectionError:
        return jsonify({
            "success": False,
            "error": "无法连接影刀，请确认影刀已运行且HTTP触发器已开启"
        }), 500
    except requests.exceptions.Timeout:
        return jsonify({"success": False, "error": "请求超时（15s）"}), 500
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/save_excel', methods=['POST'])
def save_excel():
    row_data = normalize_row_data(request.json)
    config = load_config()
    config["name"] = (row_data.get("_config_name") or config.get("name") or "闫伟杰").strip()
    try:
        save_to_excel(row_data)
        submit_result = submit_gaodun(row_data, config)
        return jsonify({
            "success": True,
            "submit": submit_result["response"],
            "display_text": submit_result["display_text"]
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


if __name__ == '__main__':
    reset_excel_file()
    print("=" * 50)
    print("  客服记录台 已启动")
    print("  本机访问: http://127.0.0.1:5000")
    print("  同事访问: http://你的电脑局域网IP:5000")
    print("  如需密码，请在 .env 中设置 APP_PASSWORD")
    print("=" * 50)
    app.run(host='0.0.0.0', port=5000, debug=False)
