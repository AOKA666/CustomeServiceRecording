import requests
import json
import openpyxl
import sys
import os


# 获取当前脚本所在目录
base_dir = os.path.dirname(os.path.abspath(__file__))

# 拼接 Excel 路径
excel_path = os.path.join(base_dir, "紧急联系人.xlsx")
cookie_path = os.path.join(base_dir, "Cookie.txt")

auth = open('C:/Users/18441/Desktop/客服记录台/Cookie.txt').read()

headers = {
    'sec-ch-ua-platform': '"Windows"',
    'x_gduserid': '97196',
    'Referer': 'https://next.gaodun.com/',
    'sec-ch-ua': '"Google Chrome";v="147", "Not.A/Brand";v="8", "Chromium";v="147"',
    'Authentication': auth,
    'sec-ch-ua-mobile': '?0',
    'X-Requested-Extend': '{"systemName":"SYSTEM-OCRM"}',
    'x_gdsid': 'no-dbU1pRhAxQ2SdP_Bvp_hACCokwMDtmtLFXI_6',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36',
    'Accept': 'application/json',
    'Content-Type': 'application/json; charset=UTF-8',
}

json_data = {
    'pageNum': 1,
    'pageSize': 10,
    'clueNo': '',
    'customerName': '',
    'customerNo': '',
    'mobile': sys.argv[1],
    'otherContact': '',
    'overseaMobile': '',
    'searchColumn': 'mobile',
}

response = requests.post('https://apigateway.gaodun.com/solon/api/v1/clues/global-quick-search', headers=headers, json=json_data)
data = json.loads(response.text)

result_list = data["result"]["list"]
target_project = sys.argv[2]

# 在列表中查找 intentProjectName == sys.argv[2] 的条目
matched = None
if len(result_list) > 1:
    for item in result_list:
        if item["intentProjectName"].strip() == target_project:
            matched = item
            break
else:
    matched = result_list[0] if result_list else None

if matched is None:
    print("查不到")
    sys.exit()

# 老师
owner_name = matched["ownerName"]
real_name = owner_name.split("-")[-1]

# 项目
intent_project = matched["intentProjectName"].strip()

# 如果 real_name 为空 或者 等于"新海"，走Excel查找
if not real_name or real_name == "新海":
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        emergency_name = None
        for row in range(1, ws.max_row + 1):
            a_value = ws.cell(row=row, column=1).value
            if a_value and str(a_value).strip() == intent_project:
                emergency_name = ws.cell(row=row, column=2).value
                break

        if emergency_name:
            real_name = str(emergency_name).strip()
        else:
            real_name = "查不到"

    except Exception as e:
        print(f"读取Excel失败：{str(e)}")

print(real_name)